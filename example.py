import json
from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('fortune_teller.xlsx')
ws = wb.active

# Load the JSON data
with open('prediction.json') as f:
    json_data = json.load(f)

# Create a dictionary for faster lookup
predictions_dict = {prediction['id']: prediction['name'] for prediction in json_data['predictions']}

# Add header for the new column
ws.cell(row=1, column=ws.max_column + 1, value='prediction_text')

# Iterate through rows and add prediction_text
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    id_cell = row[3]  # Assuming id is in the first column
    prediction_id = id_cell.value
    print(f"Processing row {id_cell.row}, id: {prediction_id}")
    prediction_text = predictions_dict.get(prediction_id, '')  # Get corresponding name from JSON data
    print(f"Prediction text: {prediction_text}")
    ws.cell(row=id_cell.row, column=ws.max_column, value=prediction_text)

# Save the modified Excel file
wb.save('fortune_teller_with_prediction.xlsx')
