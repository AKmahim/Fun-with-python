import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Data for the Excel
data = {
    "Division": ["Chittagong", "", "Dhaka", "", "", "", "", "Rajshahi", "", "Rangpur", ""],
    "Team No. & District": ["T-01 Chittagong", "T-02 B. Baria", "T-03 Gazipur", "T-04 Narayanganj", "T-05 Natara", "T-06 Madaripur", "T-07 Faridpur", "T-08 Boga", "T-09 Rajshahi", "T-10 Dinajpur", "T-11 Rangpur"],
    "Teams": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    "FY Plan": [2808, 2808, 2808, 2808, 2808, 2808, 2808, 2808, 2808, 2808, 2808],
    "YTD Plan": [48, 48, 48, 48, 48, 48, 48, 48, 48, 48, 48],
    "YTD Ach": [35, 35, 48, 48, 48, 48, 48, 48, 48, 48, 48],
    "Total %": [2, 2, 3, 3, 3, 3, 3, 3, 3, 3, 3],
    "FY Contact Plan": [56160, 56160, 112320, 112320, 112320, 112320, 112320, 112320, 112320, 112320, 112320],
    "YTD Contact Plan": [1920, 1920, 3840, 3840, 3840, 3840, 3840, 3840, 3840, 3840, 3840],
    "YTD Contact Ach": [1354, 1354, 2917, 3080, 1808, 2000, 1899, 3080, 1818, 1427, 3057],
    "YTD Contact Avg": [40.12, 40.12, 12.24, 31.79, 37.67, 41.67, 35.37, 31.98, 11.48, 29.73, 34.71],
    "FY Productive Call Plan": [18352, 18352, 36504, 36504, 36504, 36504, 36504, 36504, 36504, 36504, 36504],
    "YTD Productive Call Plan": [624, 624, 1248, 1248, 1248, 1248, 1248, 1248, 1248, 1248, 1248],
    "YTD Productive Call Ach": [482, 482, 1054, 1099, 1024, 896, 992, 1553, 1079, 745, 745],
    "No of MHW Sampling": [125, 125, 269, 219, 203, 156, 256, 419, 140, 254, 254],
    "YTD Productive Call Avg": [15.77, 15.77, 15.34, 17.35, 21.33, 18.67, 21.33, 51.88, 11.49, 16.87, 16.87],
    "YTD Sales Ach": [372, 372, 1042, 1099, 1038, 852, 992, 2383, 1097, 748, 748],
    "YTD Sales Avg": [13.77, 13.77, 20.05, 22.90, 21.62, 17.75, 20.66, 59.57, 11.49, 10.83, 10.83]
}

# Create a DataFrame
df = pd.DataFrame(data)

# Calculate Totals
totals = {
    "Division": "Grand total",
    "Team No. & District": "",
    "Teams": df["Teams"].sum(),
    "FY Plan": df["FY Plan"].sum(),
    "YTD Plan": df["YTD Plan"].sum(),
    "YTD Ach": df["YTD Ach"].sum(),
    "Total %": df["Total %"].mean(),  # Assuming the percentage is the mean of the individual percentages
    "FY Contact Plan": df["FY Contact Plan"].sum(),
    "YTD Contact Plan": df["YTD Contact Plan"].sum(),
    "YTD Contact Ach": df["YTD Contact Ach"].sum(),
    "YTD Contact Avg": df["YTD Contact Avg"].mean(),  # Assuming the average contact is the mean of the individual averages
    "FY Productive Call Plan": df["FY Productive Call Plan"].sum(),
    "YTD Productive Call Plan": df["YTD Productive Call Plan"].sum(),
    "YTD Productive Call Ach": df["YTD Productive Call Ach"].sum(),
    "No of MHW Sampling": df["No of MHW Sampling"].sum(),
    "YTD Productive Call Avg": df["YTD Productive Call Avg"].mean(),  # Assuming the average productive call is the mean of the individual averages
    "YTD Sales Ach": df["YTD Sales Ach"].sum(),
    "YTD Sales Avg": df["YTD Sales Avg"].mean()  # Assuming the average sales is the mean of the individual averages
}

# Convert totals to a DataFrame
totals_df = pd.DataFrame([totals])

# Append the totals to the original DataFrame
df = pd.concat([df, totals_df], ignore_index=True)

# Initialize a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Set the title for the sheet
ws.title = "Summary Report"

# Define header
header = ["Division", "Team No. & District", "Teams", "FY Plan", "YTD Plan", "YTD Ach", "Total %", "FY Contact Plan", "YTD Contact Plan", "YTD Contact Ach", "YTD Contact Avg", "FY Productive Call Plan", "YTD Productive Call Plan", "YTD Productive Call Ach", "No of MHW Sampling", "YTD Productive Call Avg", "YTD Sales Ach", "YTD Sales Avg"]
ws.append(header)

# Apply styles to header
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Fill data into the sheet
for row in df.itertuples(index=False, name=None):
    ws.append(row)

# Set column widths for better readability
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Apply bold font to the totals row
for cell in ws[ws.max_row]:
    cell.font = Font(bold=True)

# Save the workbook
wb.save("summary_report_with_totals.xlsx")
